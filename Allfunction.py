import openpyxl

def contriws(file,sheetName):
    wrkb=openpyxl.load_workbook(file)
    sheet=wrkb[sheetName]
    return (sheet.max_row)

def contcolumns(file,sheetName):
    wrkb=openpyxl.load_workbook(file)
    sheet=wrkb[sheetName]
    return(sheet.max_column)
def UNPWD(file,sheetName,rownum,columnno):
    wrkb=openpyxl.load_workbook(file)
    sheet=wrkb[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def result(file,sheetName,rownum,columnno,data):
    wrkb=openpyxl.load_workbook(file)
    sheet=wrkb[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    wrkb.save(file)





